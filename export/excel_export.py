"""
export/excel_export.py

Сохранение модели в Excel-файл (.xlsx).

Структура файла (4 листа):
  1_Типы деталей  — все типы деталей участка
  2_Блоки         — все блоки с параметрами
  3_Связи         — все связи (обратные выделены цветом)
  4_Маршруты      — маршрутные карты, построенные автоматически
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from model import Model, BlockType, build_routes


# ── Цвета ─────────────────────────────────────────────────────────────────────
_HDR      = PatternFill("solid", start_color="2F5496")  # тёмно-синий (шапка)
_BACK     = PatternFill("solid", start_color="FDEBD0")  # оранжевый  (обр. связь)
_ROW_ODD  = PatternFill("solid", start_color="DCE6F1")  # светло-синий
_ROW_EVEN = PatternFill("solid", start_color="FFFFFF")  # белый

_TYPE_CLR = {
    BlockType.SOURCE:    "E2EFDA",
    BlockType.TRANSPORT: "EBF3FB",
    BlockType.PROCESS:   "DCE6F1",
    BlockType.ASSEMBLY:  "E8D5F5",
    BlockType.CONTROL:   "FFF2CC",
    BlockType.BUFFER:    "F1EFE8",
    BlockType.SINK:      "F1EFE8",
}

_FONT_HDR  = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
_FONT_BOLD = Font(name="Arial", bold=True,  size=10)
_FONT_NORM = Font(name="Arial", bold=False, size=10)
_FONT_BACK = Font(name="Arial", bold=True,  color="843C0C", size=10)

_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _border():
    s = Side(style="thin", color="B8B8B8")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text):
    cell.value = text
    cell.fill, cell.font, cell.alignment, cell.border = _HDR, _FONT_HDR, _CTR, _border()

def _cell(cell, value, bold=False, fill=None, align=_LEFT, color=None):
    cell.value = value
    cell.font  = Font(name="Arial", bold=bold, size=10,
                      color=color if color else "000000")
    cell.alignment = align
    cell.border    = _border()
    if fill:
        cell.fill = fill

def _col(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ══════════════════════════════════════════════════════════════════════════════
# Главная функция
# ══════════════════════════════════════════════════════════════════════════════

def export_excel(model: Model, path: str | Path) -> None:
    """
    Создаёт Excel-файл с описанием модели производственного участка.

    Параметры:
        model — объект Model
        path  — путь к файлу (.xlsx)
    """
    wb = Workbook()
    wb.remove(wb.active)  # убираем пустой лист по умолчанию

    _sheet_details(wb, model)
    _sheet_blocks(wb, model)
    _sheet_edges(wb, model)
    _sheet_routes(wb, model)

    wb.save(Path(path))


# ── Лист 1: Типы деталей ──────────────────────────────────────────────────────

def _sheet_details(wb: Workbook, model: Model) -> None:
    ws = wb.create_sheet("1_Типы деталей")
    ws.sheet_view.showGridLines = False

    headers = ["ID детали", "Наименование", "Единица измерения", "Примечание"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(1, col), h)
    ws.row_dimensions[1].height = 22

    for row, (det_id, det) in enumerate(model.detail_types.items(), 2):
        fill = _ROW_ODD if row % 2 == 1 else _ROW_EVEN
        _cell(ws.cell(row, 1), det_id,                   bold=True, fill=fill)
        _cell(ws.cell(row, 2), det.get("label", ""),     fill=fill)
        _cell(ws.cell(row, 3), det.get("unit", "шт"),    fill=fill, align=_CTR)
        _cell(ws.cell(row, 4), det.get("note", ""),      fill=fill)
        ws.row_dimensions[row].height = 20

    for col, w in enumerate([18, 32, 20, 30], 1):
        _col(ws, col, w)
    ws.freeze_panes = "A2"


# ── Лист 2: Блоки ─────────────────────────────────────────────────────────────

def _sheet_blocks(wb: Workbook, model: Model) -> None:
    ws = wb.create_sheet("2_Блоки")
    ws.sheet_view.showGridLines = False

    headers = ["ID", "Тип", "Наименование", "Параметры", "Порты (вход)", "Порты (выход)"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(1, col), h)
    ws.row_dimensions[1].height = 22

    for row, block in enumerate(model.blocks.values(), 2):
        fill = PatternFill("solid", start_color=_TYPE_CLR.get(block.type, "FFFFFF"))

        # Параметры — компактно: "ключ=значение; ..."
        params_str = "; ".join(f"{k}={v}" for k, v in block.params.items()
                               if k not in ("input_type", "output_type"))

        _cell(ws.cell(row, 1), block.id,              bold=True, fill=fill, align=_CTR)
        _cell(ws.cell(row, 2), block.type.value,      fill=fill, align=_CTR)
        _cell(ws.cell(row, 3), block.label,           bold=True, fill=fill)
        _cell(ws.cell(row, 4), params_str,            fill=fill)
        _cell(ws.cell(row, 5), ", ".join(block.ports.in_ports),  fill=fill, align=_CTR)
        _cell(ws.cell(row, 6), ", ".join(block.ports.out_ports), fill=fill, align=_CTR)
        ws.row_dimensions[row].height = 20

    for col, w in enumerate([8, 14, 28, 38, 18, 24], 1):
        _col(ws, col, w)
    ws.freeze_panes = "A2"


# ── Лист 3: Связи ─────────────────────────────────────────────────────────────

def _sheet_edges(wb: Workbook, model: Model) -> None:
    ws = wb.create_sheet("3_Связи")
    ws.sheet_view.showGridLines = False

    headers = ["ID", "Откуда", "Куда", "Тип детали",
               "Обратная?", "Исправимость", "Время ремонта (сек)", "Комментарий"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(1, col), h)
    ws.row_dimensions[1].height = 22

    for row, edge in enumerate(model.edges.values(), 2):
        is_back = edge.is_back_edge
        fill    = _BACK if is_back else (_ROW_ODD if row % 2 == 1 else _ROW_EVEN)
        color   = "843C0C" if is_back else "000000"

        _cell(ws.cell(row, 1), edge.id,                              fill=fill, align=_CTR, color=color)
        _cell(ws.cell(row, 2), f"{edge.from_block}.{edge.from_port}",fill=fill, color=color)
        _cell(ws.cell(row, 3), f"{edge.to_block}.{edge.to_port}",    fill=fill, color=color)
        _cell(ws.cell(row, 4), edge.detail_type,                     fill=fill, align=_CTR)
        _cell(ws.cell(row, 5), "да ↻" if is_back else "—",          fill=fill, align=_CTR,
              bold=is_back, color=color)
        _cell(ws.cell(row, 6), edge.defect_type if is_back else "—", fill=fill, align=_CTR)
        _cell(ws.cell(row, 7), edge.repair_time_sec if is_back else "—", fill=fill, align=_CTR)
        _cell(ws.cell(row, 8), edge.comment,                         fill=fill)
        ws.row_dimensions[row].height = 20

    for col, w in enumerate([8, 20, 20, 14, 12, 14, 20, 28], 1):
        _col(ws, col, w)
    ws.freeze_panes = "A2"


# ── Лист 4: Маршруты ──────────────────────────────────────────────────────────

def _sheet_routes(wb: Workbook, model: Model) -> None:
    ws = wb.create_sheet("4_Маршруты")
    ws.sheet_view.showGridLines = False

    routes = build_routes(model)
    row = 1

    for route in routes:
        # Заголовок маршрута
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row, 1)
        c.value     = f"Маршрут {route.route_id}  [{route.detail_family}]  — источник: {route.source_block}"
        c.font      = Font(name="Arial", bold=True, size=11, color="2F5496")
        c.alignment = _LEFT
        ws.row_dimensions[row].height = 22
        row += 1

        # Шапка таблицы позиций
        step_headers = ["Позиция", "Блок", "Наименование операции",
                        "Тип входа", "Тип выхода", "Повтор?", "Примечание"]
        for col, h in enumerate(step_headers, 1):
            _hdr(ws.cell(row, col), h)
        ws.row_dimensions[row].height = 20
        row += 1

        # Позиции маршрута
        for step in route.steps:
            fill = _ROW_ODD if row % 2 == 0 else _ROW_EVEN
            repeat_mark = "да" if step.can_repeat else "—"
            note = "возврат брака возможен" if step.can_repeat else ""

            _cell(ws.cell(row, 1), step.position,    bold=True, fill=fill, align=_CTR)
            _cell(ws.cell(row, 2), step.block_id,    fill=fill, align=_CTR)
            _cell(ws.cell(row, 3), step.block_label, bold=True, fill=fill)
            _cell(ws.cell(row, 4), step.input_type,  fill=fill, align=_CTR)
            _cell(ws.cell(row, 5), step.output_type, fill=fill, align=_CTR)
            _cell(ws.cell(row, 6), repeat_mark,      fill=fill, align=_CTR,
                  bold=step.can_repeat,
                  color="843C0C" if step.can_repeat else "000000")
            _cell(ws.cell(row, 7), note,             fill=fill)
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1  # пустая строка между маршрутами

    for col, w in enumerate([10, 10, 30, 16, 16, 10, 28], 1):
        _col(ws, col, w)
    ws.freeze_panes = "A1"
