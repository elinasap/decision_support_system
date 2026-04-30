"""
export/excel_import.py

Загрузка модели из Excel-файла, созданного export_excel().

Читает три листа:
  1_Типы деталей  — detail_types
  2_Блоки         — blocks (ID, тип, имя, параметры, порты)
  3_Связи         — edges  (откуда/куда, обратная, исправимость, время ремонта)

Лист 4_Маршруты пропускается — маршруты строятся автоматически из модели.
"""

from pathlib import Path
from openpyxl import load_workbook
from model import Model


def import_excel(path: str | Path) -> Model:
    """
    Читает Excel-файл и восстанавливает объект Model.

    Параметры:
        path — путь к .xlsx файлу

    Возвращает:
        Model
    """
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)

    detail_types = _read_details(wb)
    blocks       = _read_blocks(wb)
    edges, used  = _read_edges(wb)

    # Проставляем использованные порты
    for bid, block_data in blocks.items():
        block_data["ports"]["used"] = list(used.get(bid, set()))

    data = {
        "name":         path.stem,
        "detail_types": detail_types,
        "blocks":       blocks,
        "edges":        edges,
    }
    wb.close()
    return Model.from_dict(data)


# ── Лист 1 ────────────────────────────────────────────────────────────────────

def _read_details(wb) -> dict:
    ws = _sheet(wb, "1_Типы деталей")
    if ws is None:
        return {}

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        det_id     = str(row[0]).strip()
        label      = str(row[1] or "").strip()
        unit       = str(row[2] or "шт").strip()
        max_repair = int(row[3]) if row[3] is not None else 0
        note       = str(row[4] or "").strip()
        if det_id:
            result[det_id] = {
                "label":      label,
                "unit":       unit,
                "max_repair": max_repair,
                "note":       note,
            }
    return result


# ── Лист 2 ────────────────────────────────────────────────────────────────────

def _read_blocks(wb) -> dict:
    ws = _sheet(wb, "2_Блоки")
    if ws is None:
        return {}

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        bid       = str(row[0]).strip()
        btype     = str(row[1] or "").strip()
        label     = str(row[2] or "").strip()
        params    = _parse_params(row[3])
        in_ports  = _split_ports(row[4])
        out_ports = _split_ports(row[5])

        if not bid:
            continue
        result[bid] = {
            "id":     bid,
            "type":   btype,
            "label":  label,
            "params": params,
            "ports":  {
                "in":   in_ports,
                "out":  out_ports,
                "used": [],
            },
        }
    return result


def _parse_params(raw) -> dict:
    """
    Разбирает строку параметров вида "key=val; key2=val2" в словарь.
    Числовые значения преобразуются в int/float.
    """
    if not raw:
        return {}
    params = {}
    for part in str(raw).split(";"):
        part = part.strip()
        if "=" not in part:
            continue
        k, v = part.split("=", 1)
        k, v = k.strip(), v.strip()
        try:
            v = int(v)
        except ValueError:
            try:
                v = float(v)
            except ValueError:
                pass
        params[k] = v
    return params


def _split_ports(raw) -> list:
    if not raw:
        return []
    return [p.strip() for p in str(raw).split(",") if p.strip()]


# ── Лист 3 ────────────────────────────────────────────────────────────────────

def _read_edges(wb) -> tuple[dict, dict]:
    """
    Возвращает (edges_dict, used_ports_dict).
    used_ports_dict: {block_id: set(port_names)}
    """
    ws = _sheet(wb, "3_Связи")
    if ws is None:
        return {}, {}

    edges = {}
    used  = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        eid      = str(row[0]).strip()
        from_str = str(row[1] or "").strip()
        to_str   = str(row[2] or "").strip()
        detail   = str(row[3] or "").strip()
        back_str = str(row[4] or "").strip()
        defect   = str(row[5] or "").strip()
        repair   = row[6]
        comment  = str(row[7] or "").strip()

        if not eid or "." not in from_str or "." not in to_str:
            continue

        from_block, from_port = from_str.rsplit(".", 1)
        to_block,   to_port   = to_str.rsplit(".", 1)

        is_back = back_str.startswith("да")

        defect_type = defect if defect not in ("—", "") else "repairable"

        try:
            repair_sec = float(repair) if repair is not None and str(repair) != "—" else 0.0
        except (ValueError, TypeError):
            repair_sec = 0.0

        edges[eid] = {
            "id":              eid,
            "from_block":      from_block,
            "from_port":       from_port,
            "to_block":        to_block,
            "to_port":         to_port,
            "is_back_edge":    is_back,
            "detail_type":     detail,
            "defect_type":     defect_type,
            "repair_time_sec": repair_sec,
            "comment":         comment,
        }

        used.setdefault(from_block, set()).add(from_port)
        used.setdefault(to_block, set()).add(to_port)

    return edges, used


# ── Вспомогательное ────────────────────────────────────────────────────────────

def _sheet(wb, name: str):
    """Возвращает лист по имени или None если не найден."""
    return wb[name] if name in wb.sheetnames else None
